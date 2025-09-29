#!/usr/bin/env python3
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE = "data.xlsx"
OUTPUT_FILE = "tool_tracker_migrated.xlsx"

HEADERS = [
    "Platform/Tool",
    "GTM Status", "GTM Notes",
    "GA4 Status", "GA4 Notes",
    "MSA Status", "MSA Notes",
    "Docs Links",
    "Example Sites",
    "WCS Team Considerations",
    "Ops Notes",
    "SK Recommended"
]

# Header colors
header_colors = {
    "Platform/Tool": "2563EB",
    "GTM Status": "DC2626", "GTM Notes": "FECACA",
    "GA4 Status": "4338CA", "GA4 Notes": "E0E7FF",
    "MSA Status": "B45309", "MSA Notes": "FED7AA",
    "Docs Links": "0F766E", "Example Sites": "99F6E4",
    "WCS Team Considerations": "6B7280", "Ops Notes": "9CA3AF",
    "SK Recommended": "065F46"
}

wrap_alignment = Alignment(wrapText=True, vertical="top", horizontal="left")

def safe_title(title: str) -> str:
    """Sanitize worksheet titles for Excel."""
    invalid = ['/', '\\', '?', '*', '[', ']', ':']
    cleaned = title
    for ch in invalid:
        cleaned = cleaned.replace(ch, '-')
    return cleaned

def style_headers(ws):
    for col, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=header)
        fill = PatternFill(start_color=header_colors.get(header,"111111"),
                           end_color=header_colors.get(header,"111111"),
                           fill_type="solid")
        c.fill = fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 28
    ws.freeze_panes = "A2"

def normalize_status(val):
    """Force messy inputs into canonical statuses."""
    if not val:
        return "Unknown"
    v = str(val).strip().lower()

    if v in ["y", "yes", "true", "1"]:
        return "Yes"
    if v in ["n", "no", "false", "0", "ntra", "not trackable"]:
        return "No"
    if "y & n" in v or "y/n" in v or "partial" in v or "maybe" in v:
        return "Partial"
    if "special" in v or "requires" in v:
        return "Special"
    if v == "?":
        return "Unknown"

    # Default: Special with nuance expected in Notes
    return "Special"

def migrate():
    wb_old = openpyxl.load_workbook(INPUT_FILE)
    wb_new = Workbook()
    first = True

    for sheet_name in wb_old.sheetnames:
        ws_old = wb_old[sheet_name]

        # Create target sheet
        if first:
            ws_new = wb_new.active
            ws_new.title = safe_title(sheet_name)
            first = False
        else:
            ws_new = wb_new.create_sheet(title=safe_title(sheet_name))

        style_headers(ws_new)
        row_cursor = 2

        # Start at row 7 (after headers)
        for r in range(7, ws_old.max_row+1):
            platform = ws_old.cell(r,1).value
            if not platform:
                continue

            # Extract values + hover comments
            gtm_val = ws_old.cell(r,2).value
            gtm_note = ws_old.cell(r,2).comment.text if ws_old.cell(r,2).comment else ""
            ga4_val = ws_old.cell(r,3).value
            ga4_note = ws_old.cell(r,3).comment.text if ws_old.cell(r,3).comment else ""
            msa_val = ws_old.cell(r,5).value
            msa_note = ws_old.cell(r,5).comment.text if ws_old.cell(r,5).comment else ""

            docs = ws_old.cell(r,4).value
            example = ws_old.cell(r,6).value
            wcs = ws_old.cell(r,7).value

            rec = [
                str(platform).strip(),

                # GTM status & notes
                normalize_status(gtm_val),
                f"{str(gtm_val).strip() if gtm_val else ''} {('| ' + gtm_note) if gtm_note else ''}".strip(),

                # GA4 status & notes
                normalize_status(ga4_val),
                f"{str(ga4_val).strip() if ga4_val else ''} {('| ' + ga4_note) if ga4_note else ''}".strip(),

                # MSA status & notes
                normalize_status(msa_val),
                f"{str(msa_val).strip() if msa_val else ''} {('| ' + msa_note) if msa_note else ''}".strip(),

                str(docs).strip() if docs else "",
                str(example).strip() if example else "",
                str(wcs).strip() if wcs else "",
                "",       # Ops Notes
                "FALSE"   # SK Recommended default
            ]

            # Write row into new sheet
            for c, val in enumerate(rec, start=1):
                cell = ws_new.cell(row=row_cursor, column=c, value=val)
                # Notes columns wrap
                if HEADERS[c-1].endswith("Notes") or HEADERS[c-1] in ["WCS Team Considerations","Ops Notes"]:
                    cell.alignment = wrap_alignment

            row_cursor += 1

    wb_new.save(OUTPUT_FILE)
    print(f"✅ Migrated workbook saved to {OUTPUT_FILE} with sanitized statuses + preserved notes")

if __name__ == "__main__":
    migrate()